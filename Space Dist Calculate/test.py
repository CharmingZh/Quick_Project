import openpyxl
from math import radians, cos, sin, asin, sqrt

"""
bridge = openpyxl.load_workbook("data/bridge.xlsx")
bus = openpyxl.load_workbook("data/bus.xlsx")
hospital = openpyxl.load_workbook("data/hospital.xlsx")
shop = openpyxl.load_workbook("data/shop.xlsx")
"""


def geodistance(lng1, lat1, lng2, lat2):
    # lng1, lat1, lng2, lat2 = (120.12802999999997, 30.28708, 115.86572000000001, 28.7427)
    lng1, lat1, lng2, lat2 = map(radians, [float(lng1), float(lat1), float(lng2), float(lat2)])  # 经纬度转换成弧度
    dlon = lng2 - lng1
    dlat = lat2 - lat1
    a = sin(dlat / 2) ** 2 + cos(lat1) * cos(lat2) * sin(dlon / 2) ** 2
    distance = 2 * asin(sqrt(a)) * 6371 * 1000  # 地球平均半径，6371km
    distance = round(distance / 1000, 3)
    return distance


def cal(file_1, file_2):
    result_wb = openpyxl.Workbook()
    result_ws = result_wb.create_sheet("Sheet", 0)

    result_wb.save("Result.xlsx")
    result_wb.close()


community = openpyxl.load_workbook("data/com.xlsx")
sheet = community['Sheet1']

row = sheet.max_row
col = sheet.max_column
print(row)
print(col)

for i in range(1, row + 1):
    for j in range(1, col + 1):
        print(sheet.cell(i, j).value, end=" ")
        result_ws.cell(i, j).value = sheet.cell(i, j).value
    print()

community.close()
